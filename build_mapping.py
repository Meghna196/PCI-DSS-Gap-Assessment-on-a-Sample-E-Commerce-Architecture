from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = Workbook()
ws = wb.active
ws.title = "PCI Mapping"

headers = ["Req ID", "Requirement Summary", "Applicable Components", "Evidence Available?", "Notes"]
ws.append(headers)

rows = [
    ["1.1", "Establish and maintain processes/documentation for network security controls (NSCs)",
     "AWS VPC, Security Groups, NACLs", "No",
     "No documented NSC ruleset review process found for ShopFast's AWS environment."],
    ["1.3", "Network access controls restrict inbound/outbound traffic to the CDE",
     "VPC, Lambda, RDS, API Gateway", "No",
     "No segmentation documented between CDE and rest of AWS VPC; Lambda and RDS share network space with non-CDE systems."],
    ["2.2", "Secure configurations are applied to all system components (no vendor defaults, hardened baselines)",
     "Lambda runtime config, RDS parameter groups, API Gateway", "Partial",
     "No evidence of a documented hardening standard or configuration baseline; default RDS parameter group appears in use (see GAP-008)."],
    ["3.3", "Sensitive Authentication Data (SAD) is not retained after authorization; account data is minimized",
     "RDS orders table (card_number column), S3 transaction logs", "No",
     "Full PANs stored at rest in card_number column; raw card data also appears in S3 logs via query strings."],
    ["3.5", "PAN is rendered unreadable anywhere it is stored (encryption, tokenization, truncation, hashing)",
     "RDS orders table", "No",
     "No encryption-at-rest or tokenization confirmed for the card_number column."],
    ["4.2", "PAN is protected with strong cryptography during transmission over open, public networks and internal untrusted networks",
     "Browser-to-frontend HTTPS, Lambda-to-RDS connection", "Partial",
     "Frontend uses HTTPS (browser leg covered). Internal Lambda-to-RDS traffic is unencrypted; PAN traverses VPC in cleartext."],
    ["5.2", "Anti-malware/malicious software protections are deployed where applicable",
     "N/A for Lambda (serverless, no persistent OS to manage); RDS underlying host managed by AWS", "N/A",
     "Requirement shifts largely to AWS's shared responsibility model for serverless compute and managed database hosts; covered under AWS's own PCI DSS AOC. No customer-managed OS/EC2 instances exist in this architecture."],
    ["6.3", "Security vulnerabilities are identified and addressed via patching and secure development practices",
     "Lambda runtime/dependencies, RDS engine version, OS layer (AWS-managed)", "Partial",
     "Application-layer patching (Lambda dependencies, npm packages) is done quarterly 'when the team gets around to it' -- insufficient cadence for critical/high vulnerabilities, which PCI DSS v4.0 expects addressed within a risk-based timeframe (typically 30 days for critical)."],
    ["7.2", "Access to system components and cardholder data is restricted based on business need to know (least privilege)",
     "Admin portal, IAM roles/policies, RDS access", "No",
     "No documented least-privilege access model or periodic access review; scenario implies broad/undifferentiated access to admin functions."],
    ["8.4", "Multi-factor authentication (MFA) is implemented for all access into the CDE",
     "Admin portal", "No",
     "Admin portal protected by username and password only; no MFA implemented."],
    ["9.x", "Physical access to systems storing, processing, or transmitting cardholder data is restricted",
     "Data center facilities (AWS-owned)", "N/A",
     "ShopFast's infrastructure runs entirely on AWS-managed cloud services (Lambda, RDS, S3). Physical security of underlying data centers is AWS's responsibility and is covered under AWS's PCI DSS Attestation of Compliance (AOC). No customer-managed physical facilities are in scope."],
    ["10.4", "Audit logs are reviewed regularly to identify anomalies or suspicious activity",
     "CloudWatch Logs", "No",
     "CloudWatch logging is enabled and logs exist, but there is no review process and no alerting configured on suspicious access patterns."],
    ["11.3", "Internal and external vulnerabilities are identified and addressed via regular scanning and penetration testing",
     "API Gateway, Lambda, RDS, public-facing frontend", "No",
     "No evidence of ASV quarterly scans, internal vulnerability scans, or an annual penetration test for the ShopFast environment."],
    ["12.8", "Policies and procedures are maintained to manage PCI DSS compliance for service providers/third parties",
     "3 third-party checkout plugins", "No",
     "None of the three third-party checkout plugins have verified SAQ-A (or equivalent) compliance documentation on file."],
]
for r in rows:
    ws.append(r)

# Styling
header_fill = PatternFill("solid", start_color="1F4E78", end_color="1F4E78")
header_font = Font(name="Arial", bold=True, color="FFFFFF", size=11)
body_font = Font(name="Arial", size=10)
wrap = Alignment(wrap_text=True, vertical="top", horizontal="left")
thin = Side(style="thin", color="B7B7B7")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

evidence_fill = {
    "No": PatternFill("solid", start_color="F8CECC", end_color="F8CECC"),
    "Partial": PatternFill("solid", start_color="FFF2CC", end_color="FFF2CC"),
    "N/A": PatternFill("solid", start_color="D9D9D9", end_color="D9D9D9"),
}

for col in range(1, 6):
    c = ws.cell(row=1, column=col)
    c.font = header_font
    c.fill = header_fill
    c.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
    c.border = border

for row_idx in range(2, ws.max_row + 1):
    for col in range(1, 6):
        c = ws.cell(row=row_idx, column=col)
        c.font = body_font
        c.alignment = wrap
        c.border = border
    ev_cell = ws.cell(row=row_idx, column=4)
    if ev_cell.value in evidence_fill:
        ev_cell.fill = evidence_fill[ev_cell.value]
        ev_cell.alignment = Alignment(wrap_text=True, vertical="top", horizontal="center")

widths = [10, 42, 32, 16, 55]
for i, w in enumerate(widths, start=1):
    ws.column_dimensions[get_column_letter(i)].width = w

ws.freeze_panes = "A2"
ws.row_dimensions[1].height = 30

wb.save("shopfast-pci-mapping.xlsx")
print("saved shopfast-pci-mapping.xlsx with", ws.max_row - 1, "requirement rows")
