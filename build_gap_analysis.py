from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = Workbook()
ws = wb.active
ws.title = "Gap Analysis"

headers = ["Gap ID", "Req ID", "Gap Description", "Risk Rating", "Likelihood", "Impact",
           "Remediation Effort", "Owner", "Target Date"]
ws.append(headers)

rows = [
    ["GAP-001", "3.3 / 3.5",
     "Full Primary Account Numbers (PANs) are stored unencrypted in the card_number column of the RDS orders table, and raw card data appears in S3 transaction logs via query string parameters.",
     "Critical", "High", "High",
     "High (schema changes, tokenization implementation, log scrubbing, data purge)",
     "Engineering Lead", "30 days"],
    ["GAP-002", "4.2",
     "Internal traffic between AWS Lambda and the RDS PostgreSQL instance does not enforce TLS, meaning PANs traversing this path are transmitted in cleartext within the VPC.",
     "High", "Medium", "High",
     "Low (enable ssl=require in connection string, enforce SSL on RDS parameter group)",
     "Engineering Lead", "14 days"],
    ["GAP-003", "1.3",
     "No network segmentation exists between the CDE (Lambda, RDS, S3 log path) and the rest of the AWS VPC, expanding the blast radius of any compromise.",
     "High", "Medium", "High",
     "Medium (define CDE subnet boundaries, security groups, NACLs; re-architect VPC routing)",
     "Cloud Infrastructure Lead", "30 days"],
    ["GAP-004", "8.4",
     "The admin portal is protected by username and password only; no multi-factor authentication is enforced for access into the CDE.",
     "High", "Medium", "High",
     "Low (enable MFA via IAM/Cognito or SSO provider, enforce via policy)",
     "Security Lead", "14 days"],
    ["GAP-005", "11.3",
     "No evidence of regular internal/external vulnerability scanning (ASV) or an annual penetration test covering the API Gateway, Lambda, or RDS.",
     "High", "Medium", "High",
     "Medium (engage ASV for quarterly scans, schedule annual pentest, remediate findings)",
     "Security Lead", "45 days"],
    ["GAP-006", "6.3",
     "OS and dependency patching is performed roughly quarterly on an ad hoc basis, well outside PCI DSS risk-based patching timelines for critical/high vulnerabilities.",
     "Medium", "Medium", "Medium",
     "Medium (adopt automated dependency scanning, define SLA-based patch cadence, e.g. 30 days for critical)",
     "Engineering Lead", "60 days"],
    ["GAP-007", "10.4",
     "CloudWatch logs are collected but never reviewed, and no alerting exists for suspicious access patterns (e.g., repeated failed logins, unusual DB queries).",
     "Medium", "Medium", "Medium",
     "Medium (configure CloudWatch alarms/metric filters, route to SIEM or on-call alerting, define review SOP)",
     "Security Operations", "45 days"],
    ["GAP-008", "2.2",
     "No documented secure configuration baseline exists for Lambda and RDS; default parameter groups and configurations appear to be in use rather than hardened, reviewed settings.",
     "Medium", "Medium", "Medium",
     "Medium (define hardening baseline, apply custom RDS parameter group, remove default credentials/settings)",
     "Engineering Lead", "60 days"],
    ["GAP-009", "7.2",
     "No least-privilege access control model or periodic access review is documented for admin portal users or IAM roles touching the CDE.",
     "Medium", "Medium", "Medium",
     "Medium (define role-based access matrix, apply least privilege to IAM policies, schedule quarterly access reviews)",
     "Security Lead", "60 days"],
    ["GAP-010", "12.8",
     "Three third-party checkout plugins are in use in the checkout flow with no verified SAQ-A (or equivalent) compliance documentation on file.",
     "Low", "Low", "Medium",
     "Low (request SAQ-A/AOC from each vendor, remove or replace any that cannot provide one)",
     "Vendor Management / Procurement", "90 days"],
]
for r in rows:
    ws.append(r)

header_fill = PatternFill("solid", start_color="1F4E78", end_color="1F4E78")
header_font = Font(name="Arial", bold=True, color="FFFFFF", size=11)
body_font = Font(name="Arial", size=10)
wrap = Alignment(wrap_text=True, vertical="top", horizontal="left")
center_wrap = Alignment(wrap_text=True, vertical="top", horizontal="center")
thin = Side(style="thin", color="B7B7B7")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

risk_fill = {
    "Critical": PatternFill("solid", start_color="C00000", end_color="C00000"),
    "High": PatternFill("solid", start_color="F8CECC", end_color="F8CECC"),
    "Medium": PatternFill("solid", start_color="FFF2CC", end_color="FFF2CC"),
    "Low": PatternFill("solid", start_color="D9EAD3", end_color="D9EAD3"),
}
risk_font = {
    "Critical": Font(name="Arial", size=10, bold=True, color="FFFFFF"),
    "High": Font(name="Arial", size=10),
    "Medium": Font(name="Arial", size=10),
    "Low": Font(name="Arial", size=10),
}

for col in range(1, len(headers) + 1):
    c = ws.cell(row=1, column=col)
    c.font = header_font
    c.fill = header_fill
    c.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
    c.border = border

for row_idx in range(2, ws.max_row + 1):
    for col in range(1, len(headers) + 1):
        c = ws.cell(row=row_idx, column=col)
        c.border = border
        if col in (1, 2, 4, 5, 6, 8, 9):
            c.alignment = center_wrap
        else:
            c.alignment = wrap
        c.font = body_font
    risk_cell = ws.cell(row=row_idx, column=4)
    val = risk_cell.value
    if val in risk_fill:
        risk_cell.fill = risk_fill[val]
        risk_cell.font = risk_font[val]

widths = [10, 10, 55, 11, 11, 11, 42, 22, 12]
for i, w in enumerate(widths, start=1):
    ws.column_dimensions[get_column_letter(i)].width = w

ws.freeze_panes = "A2"
ws.row_dimensions[1].height = 30

# Summary sheet
summary = wb.create_sheet("Summary")
summary["A1"] = "Risk Rating Matrix (Likelihood x Impact)"
summary["A1"].font = Font(bold=True, size=12)
matrix_headers = ["Likelihood \\ Impact", "Low", "Medium", "High"]
summary.append([])
summary.append(matrix_headers)
matrix_rows = [
    ["High", "Low", "High", "Critical"],
    ["Medium", "Low", "Medium", "High"],
    ["Low", "Low", "Low", "Low"],
]
for r in matrix_rows:
    summary.append(r)

for row in summary.iter_rows(min_row=3, max_row=6, min_col=1, max_col=4):
    for c in row:
        c.border = border
        c.font = Font(name="Arial", size=10)
        c.alignment = Alignment(horizontal="center", vertical="center")
summary["A3"].font = Font(bold=True, name="Arial", size=10)
for col_letter in ["A", "B", "C", "D"]:
    summary.column_dimensions[col_letter].width = 20

summary["A8"] = "Gap Count by Priority"
summary["A8"].font = Font(bold=True, size=12)
counts = {"Critical": 1, "High": 4, "Medium": 4, "Low": 1}
summary.append([])
summary.append(["Priority", "Count"])
for k, v in counts.items():
    summary.append([k, v])
for row in summary.iter_rows(min_row=10, max_row=14, min_col=1, max_col=2):
    for c in row:
        c.border = border
        c.font = Font(name="Arial", size=10)

wb.save("shopfast-gap-analysis.xlsx")
print("saved shopfast-gap-analysis.xlsx with", len(rows), "gaps")
